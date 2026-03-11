"""Premium Excel report generator — Djezzy branded, legible dark theme."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.analytics import AnalyticsService
from app.utils.site_utils import BAND_ORDER, BAND_COLORS

# ─── Excel Design System ───
RED = "E11326"
DARK_BG = "1A1A1A"
SURFACE = "242424"
HEADER_BG = "2D2D2D"
LIGHT_BG = "333333"

TITLE_FONT = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
DATA_FONT = Font(color="E0E0E0", size=10, name="Calibri")
DATA_FONT_BOLD = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
MUTED_FONT = Font(color="666666", size=10, name="Calibri")
MONO_FONT = Font(color="E11326", bold=True, size=11, name="Consolas")
SITE_FONT = Font(color="E11326", bold=True, size=12, name="Calibri")
FOOTER_FONT = Font(color="999999", size=9, name="Calibri")

THIN_BORDER = Border(
    bottom=Side(style='thin', color='333333'),
    left=Side(style='thin', color='2A2A2A'),
    right=Side(style='thin', color='2A2A2A')
)
HEADER_BORDER = Border(
    bottom=Side(style='medium', color='E11326'),
    left=Side(style='thin', color='2A2A2A'),
    right=Side(style='thin', color='2A2A2A')
)

def _fill(hex_color):
    return PatternFill(start_color=hex_color.lstrip('#'), end_color=hex_color.lstrip('#'), fill_type="solid")

def _center():
    return Alignment(horizontal='center', vertical='center')

def _left():
    return Alignment(horizontal='left', vertical='center')

def _auto_width(ws, max_col=20):
    """Compute column widths based on actual content."""
    for col_idx in range(1, max_col + 1):
        max_len = 8
        letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, 45))
        ws.column_dimensions[letter].width = max_len

def _add_dev_footer(ws, cr, max_col):
    """Add developer signature at the bottom of a sheet."""
    cr += 1
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=min(max_col, 14))
    cell = ws.cell(row=cr, column=1, value="Developed by Hicham Hedibi  ·  📞 0550 26 31 19  ·  0659 13 67 41  ·  ✉ hichamhedibi@gmail.com")
    cell.font = FOOTER_FONT
    cell.alignment = _center()
    cell.fill = _fill(DARK_BG)
    return cr


def generate_batch_report(site_codes):
    wb = Workbook()

    # ═══════════════════════════════════════════
    # SUMMARY SHEET
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = RED
    ws.sheet_view.showGridLines = False

    # Dark background for all cells
    for row in range(1, len(site_codes) + 10):
        for col in range(1, len(BAND_ORDER) + 3):
            ws.cell(row=row, column=col).fill = _fill(DARK_BG)

    # Title bar
    max_col = len(BAND_ORDER) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws['A1']
    title.value = f"DJEZZY BSS CONFIGURATION REPORT  —  {len(site_codes)} Sites"
    title.font = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
    title.fill = _fill(RED)
    title.alignment = _center()
    ws.row_dimensions[1].height = 42

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sub = ws['A2']
    sub.value = "Per-band sector configuration: S1/S2/S3/S4  (cells per sector)"
    sub.font = Font(color="888888", size=9, name="Calibri")
    sub.fill = _fill(SURFACE)
    sub.alignment = _center()

    # Column headers
    hr = 4
    headers = ["Site Code"] + BAND_ORDER
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = _fill(HEADER_BG)
        cell.alignment = _center()
        cell.border = HEADER_BORDER
    ws.row_dimensions[hr].height = 28
    ws.freeze_panes = f'A{hr + 1}'

    # Data rows
    for si, code in enumerate(site_codes):
        r = hr + 1 + si
        bands = AnalyticsService.get_config_display(code)
        band_map = {b["band"]: b["config_str"] for b in bands}

        # Alternate row shading
        row_bg = SURFACE if si % 2 == 0 else DARK_BG

        cell = ws.cell(row=r, column=1, value=code)
        cell.font = SITE_FONT
        cell.fill = _fill(row_bg)
        cell.alignment = _left()
        cell.border = THIN_BORDER

        for bi, band in enumerate(BAND_ORDER, 2):
            config_str = band_map.get(band, "—")
            cell = ws.cell(row=r, column=bi, value=config_str)
            if config_str != "—":
                bc = BAND_COLORS.get(band, "#888888").lstrip('#')
                cell.font = Font(color=bc, bold=True, size=11, name="Consolas")
            else:
                cell.font = MUTED_FONT
            cell.fill = _fill(row_bg)
            cell.alignment = _center()
            cell.border = THIN_BORDER

    # Footer
    _add_dev_footer(ws, hr + len(site_codes) + 2, max_col)

    # Auto-fit
    ws.column_dimensions['A'].width = 16
    for ci in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ═══════════════════════════════════════════
    # PER-SITE DETAIL SHEETS
    # ═══════════════════════════════════════════
    for code in site_codes[:50]:
        profile = AnalyticsService.get_site_profile(code)
        if not profile.get("technologies") and not profile.get("equipment"):
            continue

        ws2 = wb.create_sheet(title=code[:31])
        ws2.sheet_properties.tabColor = RED
        ws2.sheet_view.showGridLines = False

        # Pre-fill dark bg
        for row in range(1, 500):
            for col in range(1, 30):
                ws2.cell(row=row, column=col).fill = _fill(DARK_BG)

        # Site title
        ws2.merge_cells('A1:N1')
        t = ws2['A1']
        t.value = f"Site: {code}"
        t.font = Font(color="FFFFFF", bold=True, size=15, name="Calibri")
        t.fill = _fill(RED)
        t.alignment = _center()
        ws2.row_dimensions[1].height = 38

        # Config summary row
        bands = profile.get("config_bands", [])
        config_text = "   ·   ".join(f"{b['band']}: {b['config_str']}" for b in bands)
        ws2.merge_cells('A2:N2')
        cs = ws2['A2']
        cs.value = config_text
        cs.font = Font(color="CCCCCC", size=10, name="Consolas")
        cs.fill = _fill(SURFACE)
        cs.alignment = _center()
        ws2.row_dimensions[2].height = 24

        cr = 4
        max_data_col = 2

        # Technology tables
        tech_colors = {"2G": "3B82F6", "3G": "8B5CF6", "4G": "F59E0B", "5G": "10B981"}
        for tech in ['2G', '3G', '4G', '5G']:
            cells = profile.get("technologies", {}).get(tech, [])
            if not cells:
                continue

            tc = tech_colors.get(tech, "888888")

            # Section title
            ws2.cell(row=cr, column=1).value = f"▸ {tech} Cells ({len(cells)})"
            ws2.cell(row=cr, column=1).font = Font(color=tc, bold=True, size=12, name="Calibri")
            ws2.cell(row=cr, column=1).fill = _fill(DARK_BG)
            cr += 1

            # Column headers
            all_keys = [k for k in cells[0].keys() if not k.startswith('_')]
            col_headers = ['Sector', 'Band'] + all_keys
            max_data_col = max(max_data_col, len(col_headers))
            for ci, h in enumerate(col_headers, 1):
                cell = ws2.cell(row=cr, column=ci, value=h)
                cell.font = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
                cell.fill = _fill(tc)
                cell.alignment = _center()
                cell.border = HEADER_BORDER
            cr += 1

            # Data
            for di, cd in enumerate(cells):
                row_bg = SURFACE if di % 2 == 0 else DARK_BG

                s_cell = ws2.cell(row=cr, column=1, value=f"S{cd.get('_sector', '?')}")
                s_cell.font = DATA_FONT_BOLD
                s_cell.fill = _fill(row_bg)
                s_cell.alignment = _center()
                s_cell.border = THIN_BORDER

                b_cell = ws2.cell(row=cr, column=2, value=cd.get('_band', '-'))
                bc = BAND_COLORS.get(cd.get('_band', ''), '#888').lstrip('#')
                b_cell.font = Font(color=bc, bold=True, size=10, name="Calibri")
                b_cell.fill = _fill(row_bg)
                b_cell.alignment = _center()
                b_cell.border = THIN_BORDER

                for ci, k in enumerate(all_keys, 3):
                    v = cd.get(k, '')
                    if v == '-':
                        v = ''
                    cell = ws2.cell(row=cr, column=ci, value=v)
                    cell.font = DATA_FONT
                    cell.fill = _fill(row_bg)
                    cell.alignment = _left()
                    cell.border = THIN_BORDER
                cr += 1
            cr += 1  # gap between sections

        # Equipment
        equipment = profile.get("equipment", [])
        if equipment:
            ws2.cell(row=cr, column=1).value = f"▸ Equipment ({len(equipment)})"
            ws2.cell(row=cr, column=1).font = Font(color="FF6B35", bold=True, size=12, name="Calibri")
            cr += 1

            eq_keys = [k for k in equipment[0].keys() if not k.startswith('_')]
            max_data_col = max(max_data_col, len(eq_keys))
            for ci, h in enumerate(eq_keys, 1):
                cell = ws2.cell(row=cr, column=ci, value=h)
                cell.font = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
                cell.fill = _fill("FF6B35")
                cell.alignment = _center()
                cell.border = HEADER_BORDER
            cr += 1

            for di, eq in enumerate(equipment):
                row_bg = SURFACE if di % 2 == 0 else DARK_BG
                for ci, k in enumerate(eq_keys, 1):
                    v = eq.get(k, '')
                    if v == '-':
                        v = ''
                    cell = ws2.cell(row=cr, column=ci, value=v)
                    cell.font = DATA_FONT
                    cell.fill = _fill(row_bg)
                    cell.alignment = _left()
                    cell.border = THIN_BORDER
                cr += 1

        # Developer footer
        _add_dev_footer(ws2, cr + 1, max_data_col)

        # Auto-fit columns
        _auto_width(ws2, max_data_col)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
